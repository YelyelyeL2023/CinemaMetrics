import pandas as pd
import psycopg2
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import json
import os
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter

# Database configuration
DB_CONFIG = {
    'host': 'localhost',
    'database': 'Data',
    'user': 'postgres',
    'password': '2006',
    'port': 5432
}

# Create charts directory if it doesn't exist
os.makedirs('charts', exist_ok=True)
os.makedirs('exports', exist_ok=True)

# Set style for better looking charts
plt.style.use('seaborn-v0_8')
sns.set_palette("husl")

def get_db_connection():
    """Establish database connection."""
    return psycopg2.connect(**DB_CONFIG)

def execute_query(query, description):
    """Execute SQL query and return DataFrame."""
    print(f"\nExecuting query: {description}")
    print("-" * 60)
    
    conn = get_db_connection()
    df = pd.read_sql_query(query, conn)
    conn.close()
    
    print(f"Query returned {len(df)} rows")
    print(f"Columns: {list(df.columns)}")
    print(df.head())
    
    return df

def export_to_excel(dataframes_dict, filename):
    """Export DataFrames to a formatted Excel workbook."""
    safe_sheets = {}
    for sheet_name, df in dataframes_dict.items():
        base_name = sheet_name[:31]
        candidate = base_name
        counter = 1
        while candidate in safe_sheets:
            suffix = f"_{counter}"
            candidate = f"{base_name[:31-len(suffix)]}{suffix}"
            counter += 1
        safe_sheets[candidate] = df

    output_path = os.path.join('exports', filename)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, df in safe_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        for sheet_name, df in safe_sheets.items():
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = "B2"
            ws.auto_filter.ref = ws.dimensions
            max_row = ws.max_row
            if max_row <= 1:
                continue

            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if not len(numeric_cols):
                continue

            min_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
            max_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")

            for col_name in numeric_cols:
                idx = df.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(idx)
                cell_range = f"{col_letter}2:{col_letter}{max_row}"

                gradient_rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                    end_type="max", end_color="FF00AA00"
                )
                min_rule = CellIsRule(
                    operator="equal",
                    formula=[f"MIN(${col_letter}$2:${col_letter}${max_row})"],
                    fill=min_fill
                )
                max_rule = CellIsRule(
                    operator="equal",
                    formula=[f"MAX(${col_letter}$2:${col_letter}${max_row})"],
                    fill=max_fill
                )

                ws.conditional_formatting.add(cell_range, gradient_rule)
                ws.conditional_formatting.add(cell_range, min_rule)
                ws.conditional_formatting.add(cell_range, max_rule)

    total_rows = sum(len(df) for df in safe_sheets.values())
    print(f"Created file {filename}, {len(safe_sheets)} sheets, {total_rows} rows")

def create_missing_tables():
    """Create missing keywords and ratings tables if they don't exist."""
    conn = get_db_connection()
    
    try:
        with conn.cursor() as cursor:
            # Create keywords table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS keywords (
                    id INTEGER PRIMARY KEY,
                    keywords JSONB
                );
            """)
            
            # Create ratings table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS ratings (
                    userId INTEGER,
                    movieId INTEGER,
                    rating FLOAT,
                    timestamp BIGINT
                );
            """)
            
            conn.commit()
            print("✓ Created missing tables (keywords, ratings)")
    except Exception as e:
        print(f"Note: Tables might already exist or error occurred: {e}")
    finally:
        conn.close()

def create_pie_chart():
    """1. Pie Chart: Genre Distribution of High-Budget Movies (>$50M)"""
    query = """
    SELECT 
        jsonb_extract_path_text(genre, 'name') as genre_name,
        COUNT(*) as movie_count
    FROM movies_metadata m
    JOIN credits c ON m.id = c.id,
         jsonb_array_elements(m.genres) as genre
    WHERE m.budget > 50000000 
      AND m.genres IS NOT NULL
    GROUP BY jsonb_extract_path_text(genre, 'name')
    HAVING COUNT(*) >= 5
    ORDER BY movie_count DESC
    LIMIT 8;
    """
    
    df = execute_query(query, "Genre distribution of high-budget movies (>$50M)")
    
    # Create pie chart
    plt.figure(figsize=(10, 8))
    colors = plt.cm.Set3(np.linspace(0, 1, len(df)))
    wedges, texts, autotexts = plt.pie(df['movie_count'], 
                                      labels=df['genre_name'], 
                                      autopct='%1.1f%%',
                                      colors=colors,
                                      startangle=90)
    
    plt.title('Genre Distribution of High-Budget Movies (>$50M)\nBased on Movies with Credits Data', 
              fontsize=14, fontweight='bold', pad=20)
    plt.axis('equal')
    
    # Make percentage text more readable
    for autotext in autotexts:
        autotext.set_color('white')
        autotext.set_fontweight('bold')
    
    plt.tight_layout()
    plt.savefig('charts/pie_chart_genre_distribution.png', dpi=300, bbox_inches='tight')
    plt.show()
    
    print(f"✓ Created PIE CHART: Genre distribution with {len(df)} genres")
    print("Shows which genres dominate high-budget movie production")
    return df

def create_bar_chart():
    """2. Bar Chart: Average Movie Rating by Production Company (Top 15)"""
    query = """
    SELECT 
        jsonb_extract_path_text(company, 'name') as company_name,
        COUNT(*) as movie_count,
        ROUND(AVG(m.vote_average)::numeric, 2) as avg_rating,
        ROUND(AVG(m.revenue)::numeric, 0) as avg_revenue
    FROM movies_metadata m
    JOIN credits c ON m.id = c.id,
         jsonb_array_elements(m.production_companies) as company
    WHERE m.production_companies IS NOT NULL 
      AND m.vote_average > 0
    GROUP BY jsonb_extract_path_text(company, 'name')
    HAVING COUNT(*) >= 3 AND AVG(m.vote_average) > 0
    ORDER BY avg_rating DESC
    LIMIT 15;
    """
    
    df = execute_query(query, "Average movie rating by production company")
    
    # Create bar chart
    plt.figure(figsize=(12, 8))
    bars = plt.bar(range(len(df)), df['avg_rating'], 
                   color=plt.cm.viridis(np.linspace(0, 1, len(df))))
    
    plt.xlabel('Production Companies', fontsize=12, fontweight='bold')
    plt.ylabel('Average Movie Rating', fontsize=12, fontweight='bold')
    plt.title('Average Movie Rating by Production Company\n(Companies with 3+ Movies and Credits Data)', 
              fontsize=14, fontweight='bold', pad=20)
    
    # Rotate x-axis labels for better readability
    plt.xticks(range(len(df)), df['company_name'], rotation=45, ha='right')
    plt.ylim(0, 10)
    
    # Add value labels on bars
    for i, bar in enumerate(bars):
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                f'{height}', ha='center', va='bottom', fontweight='bold')
    
    plt.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    plt.savefig('charts/bar_chart_company_ratings.png', dpi=300, bbox_inches='tight')
    plt.show()
    
    print(f"✓ Created BAR CHART: Production companies with {len(df)} companies")
    print("Shows which production companies consistently create higher-rated movies")
    return df

def create_horizontal_bar_chart():
    """3. Horizontal Bar Chart: Top Actors by Number of High-Rated Movies (Rating >= 7.5)"""
    query = """
    SELECT 
        jsonb_extract_path_text(cast_member, 'name') as actor_name,
        COUNT(*) as high_rated_movies,
        ROUND(AVG(m.vote_average)::numeric, 2) as avg_movie_rating,
        STRING_AGG(DISTINCT jsonb_extract_path_text(genre, 'name'), ', ') as common_genres
    FROM movies_metadata m
    JOIN credits c ON m.id = c.id,
         jsonb_array_elements(c."cast") as cast_member,
         jsonb_array_elements(m.genres) as genre
    WHERE m.vote_average >= 7.5 
      AND m.vote_count >= 100
      AND jsonb_extract_path_text(cast_member, 'name') IS NOT NULL
    GROUP BY jsonb_extract_path_text(cast_member, 'name')
    HAVING COUNT(*) >= 3
    ORDER BY high_rated_movies DESC, avg_movie_rating DESC
    LIMIT 12;
    """
    
    df = execute_query(query, "Top actors by number of high-rated movies")
    
    # Create horizontal bar chart
    plt.figure(figsize=(12, 10))
    colors = plt.cm.plasma(np.linspace(0, 1, len(df)))
    bars = plt.barh(range(len(df)), df['high_rated_movies'], color=colors)
    
    plt.xlabel('Number of High-Rated Movies (Rating ≥ 7.5)', fontsize=12, fontweight='bold')
    plt.ylabel('Actors', fontsize=12, fontweight='bold')
    plt.title('Top Actors by Number of High-Rated Movies\n(Movies with Rating ≥ 7.5 and 100+ Votes)', 
              fontsize=14, fontweight='bold', pad=20)
    
    plt.yticks(range(len(df)), df['actor_name'])
    
    # Add value labels on bars
    for i, bar in enumerate(bars):
        width = bar.get_width()
        plt.text(width + 0.1, bar.get_y() + bar.get_height()/2.,
                f'{int(width)}', ha='left', va='center', fontweight='bold')
    
    plt.grid(axis='x', alpha=0.3)
    plt.tight_layout()
    plt.savefig('charts/horizontal_bar_chart_top_actors.png', dpi=300, bbox_inches='tight')
    plt.show()
    
    print(f"✓ Created HORIZONTAL BAR CHART: Top actors with {len(df)} actors")
    print("Shows which actors appear most frequently in critically acclaimed movies")
    return df

def create_line_chart():
    """4. Line Chart: Average Movie Budget and Revenue Trends Over Years"""
    query = """
    SELECT 
        EXTRACT(YEAR FROM m.release_date) as release_year,
        COUNT(*) as movie_count,
        ROUND(AVG(m.budget)::numeric, 0) as avg_budget,
        ROUND(AVG(m.revenue)::numeric, 0) as avg_revenue,
        ROUND(AVG(m.vote_average)::numeric, 2) as avg_critic_rating
    FROM movies_metadata m
    JOIN credits c ON m.id = c.id
    WHERE m.release_date IS NOT NULL 
      AND m.budget > 0 
      AND m.revenue > 0
      AND EXTRACT(YEAR FROM m.release_date) BETWEEN 1980 AND 2020
    GROUP BY EXTRACT(YEAR FROM m.release_date)
    HAVING COUNT(*) >= 5
    ORDER BY release_year;
    """
    
    df = execute_query(query, "Movie budget and revenue trends over years")
    
    # Create line chart with dual y-axis
    fig, ax1 = plt.subplots(figsize=(14, 8))
    
    # Plot budget and revenue
    ax1.plot(df['release_year'], df['avg_budget'], 
             marker='o', linewidth=2, label='Average Budget', color='blue')
    ax1.plot(df['release_year'], df['avg_revenue'], 
             marker='s', linewidth=2, label='Average Revenue', color='green')
    
    ax1.set_xlabel('Release Year', fontsize=12, fontweight='bold')
    ax1.set_ylabel('Amount (USD)', fontsize=12, fontweight='bold')
    ax1.set_title('Movie Budget and Revenue Trends Over Time\n(Movies with Credits Data)', 
                  fontsize=14, fontweight='bold', pad=20)
    
    # Format y-axis to show values in millions
    ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x/1e6:.0f}M'))
    
    # Create second y-axis for critic ratings
    ax2 = ax1.twinx()
    ax2.plot(df['release_year'], df['avg_critic_rating'], 
             marker='d', linewidth=2, label='Average Critic Rating', color='red', alpha=0.7)
    ax2.set_ylabel('Average Critic Rating', fontsize=12, fontweight='bold', color='red')
    ax2.set_ylim(0, 10)
    
    # Combine legends
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left')
    
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig('charts/line_chart_budget_revenue_trends.png', dpi=300, bbox_inches='tight')
    plt.show()
    
    print(f"✓ Created LINE CHART: Budget/revenue trends with {len(df)} years")
    print("Shows how movie budgets, revenues, and critic ratings evolved over time")
    return df

def create_histogram():
    """5. Histogram: Distribution of Movie Runtimes for Different Genres"""
    query = """
    SELECT 
        m.runtime,
        jsonb_extract_path_text(genre, 'name') as genre_name,
        m.vote_average
    FROM movies_metadata m
    JOIN credits c ON m.id = c.id,
         jsonb_array_elements(m.genres) as genre
    WHERE m.runtime BETWEEN 60 AND 240
      AND m.runtime IS NOT NULL
      AND jsonb_extract_path_text(genre, 'name') IN ('Action', 'Drama', 'Comedy', 'Thriller', 'Romance')
      AND m.vote_count >= 50
    ORDER BY m.runtime;
    """
    
    df = execute_query(query, "Movie runtime distribution by genre")
    
    # Create histogram
    plt.figure(figsize=(12, 8))
    
    # Create separate histograms for each genre
    genres = df['genre_name'].unique()
    colors = plt.cm.Set2(np.linspace(0, 1, len(genres)))
    
    for i, genre in enumerate(genres):
        genre_data = df[df['genre_name'] == genre]['runtime']
        plt.hist(genre_data, bins=20, alpha=0.7, label=genre, color=colors[i])
    
    plt.xlabel('Movie Runtime (minutes)', fontsize=12, fontweight='bold')
    plt.ylabel('Number of Movies', fontsize=12, fontweight='bold')
    plt.title('Distribution of Movie Runtimes by Genre\n(Movies with Credits and 50+ Votes)', 
              fontsize=14, fontweight='bold', pad=20)
    
    plt.legend()
    plt.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    plt.savefig('charts/histogram_runtime_distribution.png', dpi=300, bbox_inches='tight')
    plt.show()
    
    print(f"✓ Created HISTOGRAM: Runtime distribution with {len(df)} movies")
    print("Shows how movie lengths vary across different genres")
    return df

def create_scatter_plot():
    """6. Scatter Plot: Movie Budget vs. Revenue with Vote Average Color Coding"""
    query = """
    SELECT 
        m.budget,
        m.revenue,
        m.vote_average,
        m.title,
        COUNT(DISTINCT jsonb_extract_path_text(cast_member, 'name')) as cast_size,
        jsonb_extract_path_text(jsonb_array_elements(m.genres), 'name') as primary_genre
    FROM movies_metadata m
    JOIN credits c ON m.id = c.id,
         jsonb_array_elements(c."cast") as cast_member
    WHERE m.budget > 1000000 
      AND m.revenue > 1000000
      AND m.budget < m.revenue
      AND m.vote_count >= 100
      AND m.vote_average > 0
    GROUP BY m.id, m.budget, m.revenue, m.vote_average, m.title, m.genres
    HAVING COUNT(DISTINCT jsonb_extract_path_text(cast_member, 'name')) >= 5
    ORDER BY m.revenue DESC
    LIMIT 200;
    """
    
    df = execute_query(query, "Movie budget vs revenue with ratings")
    
    # Create scatter plot
    plt.figure(figsize=(14, 10))
    
    # Create scatter plot with color mapping based on vote average
    scatter = plt.scatter(df['budget'], df['revenue'], 
                         c=df['vote_average'], 
                         s=df['cast_size']*2,  # Size based on cast size
                         alpha=0.7, 
                         cmap='RdYlGn',
                         edgecolors='black',
                         linewidth=0.5)
    
    plt.xlabel('Budget (USD)', fontsize=12, fontweight='bold')
    plt.ylabel('Revenue (USD)', fontsize=12, fontweight='bold')
    plt.title('Movie Budget vs Revenue\n(Color: Vote Average, Size: Cast Size, with Credits Data)', 
              fontsize=14, fontweight='bold', pad=20)
    
    # Format axes to show values in millions
    plt.gca().xaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x/1e6:.0f}M'))
    plt.gca().yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x/1e6:.0f}M'))
    
    # Add colorbar
    cbar = plt.colorbar(scatter)
    cbar.set_label('Vote Average Rating', fontsize=12, fontweight='bold')
    
    # Add diagonal line showing break-even point
    max_val = max(df['budget'].max(), df['revenue'].max())
    plt.plot([0, max_val], [0, max_val], 'r--', alpha=0.5, label='Break-even line')
    
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig('charts/scatter_plot_budget_vs_revenue.png', dpi=300, bbox_inches='tight')
    plt.show()
    
    print(f"✓ Created SCATTER PLOT: Budget vs revenue with {len(df)} movies")
    print("Shows relationship between budget, revenue, vote ratings, and cast size")
    return df

def main():
    """Main function to create all visualizations."""
    print("🎬 CinemaMetrics Analytics - Creating Visualizations")
    print("=" * 60)
    
    try:
        # Test database connection
        conn = get_db_connection()
        conn.close()
        print("✓ Database connection successful")
        
        # Create missing tables (if needed)
        create_missing_tables()
        
        pie_df = create_pie_chart()
        bar_df = create_bar_chart()
        actors_df = create_horizontal_bar_chart()
        trends_df = create_line_chart()
        runtime_df = create_histogram()
        scatter_df = create_scatter_plot()

        export_to_excel(
            {
                "HighBudgetGenres": pie_df,
                "CompanyRatings": bar_df,
                "TopActorsHighRated": actors_df,
                "BudgetRevenueTrends": trends_df,
                "RuntimeDistribution": runtime_df,
                "BudgetVsRevenue": scatter_df
            },
            "cinemametrics_report.xlsx"
        )
        
        print("\n" + "=" * 60)
        print("📊 VISUALIZATION SUMMARY")
        print("=" * 60)
        print("✓ Pie Chart: Genre distribution of high-budget movies")
        print("✓ Bar Chart: Average ratings by production company")
        print("✓ Horizontal Bar: Top actors in high-rated movies")
        print("✓ Line Chart: Budget and revenue trends over time")
        print("✓ Histogram: Runtime distribution by genre")
        print("✓ Scatter Plot: Budget vs revenue with rating colors")
        print("✓ Excel Report: cinemametrics_report.xlsx generated in exports/")
        print("\nAll charts saved to 'charts/' directory")
        print("Each query uses multiple JOINs for comprehensive analysis")
        
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    main()